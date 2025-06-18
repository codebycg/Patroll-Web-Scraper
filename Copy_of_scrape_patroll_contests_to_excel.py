from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
import openpyxl
from openpyxl.utils import get_column_letter

# Load the Excel file
excel_path = "PatentPlusAI Week 2 Deliverable.xlsx"
try:
    workbook = openpyxl.load_workbook(excel_path)
except Exception as e:
    print("Error loading Excel file:", e)
    raise

# Create a new sheet for the output if it doesn't exist
sheet_name = "Scraped Contests"
if sheet_name not in workbook.sheetnames:
    ws = workbook.create_sheet(sheet_name)
    ws.append(["Troll Patent", "Prior Art Patents", "Contest Title", "Award Amount", "Contest URL"])
else:
    ws = workbook[sheet_name]

# Setup Selenium browser
options = Options()
options.add_argument("--headless")
options.add_argument("--window-size=1920,1080")
driver = webdriver.Chrome(options=options)

base_url = "https://patroll.unifiedpatents.com"
main_url = f"{base_url}/contests?category=finished"
driver.get(main_url)

max_pages = 2  # Adjust if you want more contests
try:
    for page in range(1, max_pages + 1):
        print(f"Processing contest listings page {page}...")

        time.sleep(2)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        contest_links = soup.select("ul.ant-list-items a[href^='/contests/']")
        contest_urls = list(set([base_url + a['href'] for a in contest_links]))

        for contest_url in contest_urls:
            print(f"Opening contest: {contest_url}")
            driver.get(contest_url)
            time.sleep(2)

            soup = BeautifulSoup(driver.page_source, "html.parser")

            # Contest Title
            title_elem = soup.find("h1")
            title = title_elem.text.strip() if title_elem else "N/A"

            # Troll Patent
            troll_patent = "N/A"
            troll_patent_link = soup.find("a", href=True)
            if troll_patent_link and "google.com" in troll_patent_link["href"]:
                troll_patent = troll_patent_link.text.strip()

            # Award Amount
            award_elem = soup.find("div", text=lambda t: t and "Award Amount" in t)
            award_amount = award_elem.find_next("div").text.strip() if award_elem else "N/A"

            # Prior Art Patents
            prior_art = []
            all_links = soup.find_all("a", href=True)
            for a in all_links:
                href = a["href"]
                if href.startswith("https://www.google.com/patents"):
                    prior_art.append(href.split("patents/")[-1])

            prior_art_str = ", ".join(set(prior_art))

            # Add row to Excel
            ws.append([troll_patent, prior_art_str, title, award_amount, contest_url])

        # Next Page
        try:
            next_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, f"//li[@title='{page + 1}']"))
            )
            driver.execute_script("arguments[0].scrollIntoView();", next_button)
            driver.execute_script("arguments[0].click();", next_button)
        except:
            print("No more pages.")
            break

finally:
    driver.quit()
    workbook.save(excel_path)
    print(f"\n✅ Data saved directly to '{sheet_name}' in Excel file.")
